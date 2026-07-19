from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse

from accounts.models import User, ClientProfile
from accounting.models import AccountTransaction
from staff.permissions import perm_required


def _clients_with_balance():
    """
    كل العملاء النشطين مع رصيدهم الحالي (موجب = عليه فلوس)، مرتبين من الأكتر
    مديونية للأقل. بنستخدم annotate على مستوى قاعدة البيانات (مش balance_for
    في لوب) عشان الأداء يفضل كويس حتى لو عدد العملاء كبر.
    """
    balances = dict(
        AccountTransaction.objects.values('client_id').annotate(total=Sum('amount')).values_list('client_id', 'total')
    )
    # select_related('account_type') كمان (مش user بس) — عرض المديونيات
    # وتصدير الإكسل بيوصلوا لـ profile.account_type.name لكل عميل، فمن غيرها
    # كانت هتبقى N+1 (استعلام إضافي منفصل لكل عميل نشط).
    profiles = ClientProfile.objects.filter(user__status='ACTIVE').select_related('user', 'account_type')
    rows = []
    for profile in profiles:
        balance = balances.get(profile.user_id) or Decimal('0')
        rows.append({'profile': profile, 'balance': balance, 'balance_abs': abs(balance)})
    rows.sort(key=lambda r: r['balance'], reverse=True)
    return rows


@perm_required('accounting.view_accounttransaction')
def accounting_overview(request):
    rows = _clients_with_balance()

    total_receivable = sum((r['balance'] for r in rows if r['balance'] > 0), Decimal('0'))
    total_credit = sum((-r['balance'] for r in rows if r['balance'] < 0), Decimal('0'))
    debtors_count = sum(1 for r in rows if r['balance'] > 0)

    recent_transactions = (
        AccountTransaction.objects.select_related('client', 'client__client_profile', 'invoice', 'created_by')
        .order_by('-created_at')[:50]
    )

    active_clients = [r['profile'] for r in rows]

    return render(request, 'staff/accounting/overview.html', {
        'rows': rows,
        'total_receivable': total_receivable,
        'total_credit': total_credit,
        'debtors_count': debtors_count,
        'recent_transactions': recent_transactions,
        'active_clients': active_clients,
        'payment_methods': AccountTransaction.PaymentMethod.choices,
    })


@perm_required('accounting.add_accounttransaction')
def accounting_quick_entry(request):
    if request.method != 'POST':
        return redirect('staff:accounting_overview')

    client_id = request.POST.get('client_id', '').strip()
    kind = request.POST.get('kind', '').strip()
    raw_amount = request.POST.get('amount', '').strip()
    method = request.POST.get('method', '')
    direction = request.POST.get('direction', 'increase')
    note = request.POST.get('note', '').strip()

    profile = get_object_or_404(ClientProfile, user_id=client_id, user__status='ACTIVE')

    try:
        amount = Decimal(raw_amount)
    except (InvalidOperation, TypeError):
        amount = None

    if not amount or amount <= 0:
        messages.error(request, 'يجب أن تكون القيمة رقمًا أكبر من صفر.')
        return redirect('staff:accounting_overview')

    if kind == AccountTransaction.Kind.PAYMENT:
        try:
            AccountTransaction.objects.create(
                client=profile.user,
                kind=AccountTransaction.Kind.PAYMENT,
                amount=-amount,
                method=method,
                note=note,
                created_by=request.user,
            )
        except ValidationError as e:
            messages.error(request, f'المبلغ غير صالح: {"، ".join(e.messages)}')
        else:
            messages.success(request, f'تم تسجيل دفعة بقيمة {amount} ج.م لـ {profile.business_name}.')

    elif kind == AccountTransaction.Kind.ADJUSTMENT:
        if not note:
            messages.error(request, 'يجب إدخال سبب أو ملاحظة مع عملية التسوية.')
            return redirect('staff:accounting_overview')
        signed_amount = amount if direction == 'increase' else -amount
        try:
            AccountTransaction.objects.create(
                client=profile.user,
                kind=AccountTransaction.Kind.ADJUSTMENT,
                amount=signed_amount,
                note=note,
                created_by=request.user,
            )
        except ValidationError as e:
            messages.error(request, f'المبلغ غير صالح: {"، ".join(e.messages)}')
        else:
            messages.success(request, f'تم تسجيل تسوية لـ {profile.business_name}.')
    else:
        messages.error(request, 'نوع الحركة غير معروف.')

    return redirect('staff:accounting_overview')


@perm_required('accounting.view_accounttransaction')
def accounting_export(request):
    import openpyxl
    from openpyxl.utils import get_column_letter

    rows = _clients_with_balance()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المديونيات'
    headers = ['اسم النشاط', 'نوع الحساب', 'الهاتف', 'الرصيد (ج.م)']
    ws.append(headers)
    for row in rows:
        profile = row['profile']
        ws.append([
            profile.business_name,
            profile.account_type.name,
            profile.phone,
            float(row['balance']),
        ])
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = 24

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="biozone_accounts_receivable.xlsx"'
    wb.save(response)
    return response
