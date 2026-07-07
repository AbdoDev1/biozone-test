from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import ProtectedError
from products.models import Product, ProductUnit, Category
from products.forms import ProductForm, ProductUnitForm, ProductUnitFormSet
from products.pricing import autofill_small_unit_price
from inventory.models import Inventory, StockMovement
import openpyxl


def staff_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role not in ['ADMIN', 'WAREHOUSE']:
            return redirect('staff:login')
        return view_func(request, *args, **kwargs)
    return wrapper


@staff_required
def product_list(request):
    products = Product.objects.select_related('category').prefetch_related('units').all()
    categories = Category.objects.filter(is_active=True)
    selected_category = request.GET.get('category', '')
    search_q = request.GET.get('q', '')
    if selected_category:
        products = products.filter(category__slug=selected_category)
    if search_q:
        products = products.filter(name_ar__icontains=search_q)
    return render(request, 'staff/products/list.html', {
        'products': products,
        'categories': categories,
        'selected_category': selected_category,
        'search_q': search_q,
    })


@staff_required
def product_add(request):
    if request.method == 'POST':
        # لو سعر القطعة (الوحدة الصغرى) سايبينه فاضي وفي كرتونة (وحدة كبرى)
        # بسعر وكمية، بنحسب سعر القطعة تلقائيًا قبل ما نبني الفورم/الفورمست
        post_data = autofill_small_unit_price(request.POST)
        form = ProductForm(post_data, request.FILES)
        formset = ProductUnitFormSet(post_data, instance=Product())
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    product = form.save()
                    formset.instance = product
                    units = formset.save()  # بيحفظ كل الوحدات الجديدة (النماذج اللي اتملت واتصدّق عليها)
            except IntegrityError:
                messages.error(
                    request,
                    'حدث تعارض غير متوقع أثناء حفظ وحدات المنتج (مثلاً وحدتين بنفس '
                    'الحجم) — يرجى مراجعة الوحدات وإعادة المحاولة. لو المشكلة '
                    'استمرت، يرجى إبلاغ فريق التطوير بخطوات إعادة حدوثها بالتفصيل.'
                )
                return render(request, 'staff/products/form.html', {
                    'form': form, 'formset': formset,
                    'title': 'إضافة منتج جديد', 'is_edit': False,
                })

            inventory, _ = Inventory.objects.get_or_create(
                product=product,
                defaults={'quantity': 0, 'reserved': 0, 'min_quantity': 0},
            )
            # initial_stock مش حقل موديل — بنقراه من cleaned_data لكل نموذج غير محذوف
            for unit_form in formset.forms:
                if unit_form.cleaned_data.get('DELETE'):
                    continue
                unit = unit_form.instance
                initial_stock = unit_form.cleaned_data.get('initial_stock') or 0
                if unit.pk and initial_stock > 0:
                    StockMovement.objects.create(
                        inventory=inventory,
                        unit=unit,
                        movement_type='IN',
                        quantity=initial_stock,
                        note='كمية ابتدائية عند إضافة المنتج',
                        created_by=request.user,
                    )
            messages.success(request, f'تم إضافة المنتج "{product.name_ar}" بنجاح.')
            return redirect('staff:product_list')
    else:
        form = ProductForm()
        formset = ProductUnitFormSet(instance=Product())
    return render(request, 'staff/products/form.html', {
        'form': form,
        'formset': formset,
        'title': 'إضافة منتج جديد',
        'is_edit': False,
    })


@staff_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        # نفس منطق الحساب التلقائي المستخدم في الإضافة (شوف autofill_small_unit_price)
        post_data = autofill_small_unit_price(request.POST)
        form = ProductForm(post_data, request.FILES, instance=product)
        formset = ProductUnitFormSet(post_data, instance=product)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    formset.save()
            except ProtectedError:
                messages.error(
                    request,
                    'مينفعش تمسح وحدة ليها حركات مخزون أو طلبات مسجّلة عليها — '
                    'عطّل استخدامها بدل الحذف، أو سيبها من غير حذف.'
                )
                return render(request, 'staff/products/form.html', {
                    'form': form, 'formset': formset,
                    'title': f'تعديل: {product.name_ar}', 'is_edit': True, 'product': product,
                })
            except IntegrityError:
                messages.error(
                    request,
                    'حدث تعارض غير متوقع أثناء حفظ وحدات المنتج (مثلاً وحدتين بنفس '
                    'الحجم) — يرجى إعادة تحميل الصفحة والتأكد إن كل وحدة (صغرى/كبرى) '
                    'ليها حجم مختلف عن التانية، وإعادة المحاولة. لو المشكلة استمرت، '
                    'يرجى إبلاغ فريق التطوير بخطوات إعادة حدوثها بالتفصيل.'
                )
                return render(request, 'staff/products/form.html', {
                    'form': form, 'formset': formset,
                    'title': f'تعديل: {product.name_ar}', 'is_edit': True, 'product': product,
                })

            # أي وحدة جديدة اتضافت أثناء التعديل ومعاها كمية ابتدائية
            inventory, _ = Inventory.objects.get_or_create(
                product=product,
                defaults={'quantity': 0, 'reserved': 0, 'min_quantity': 0},
            )
            for unit_form in formset.forms:
                if unit_form.cleaned_data.get('DELETE'):
                    continue
                initial_stock = unit_form.cleaned_data.get('initial_stock') or 0
                unit = unit_form.instance
                if unit.pk and initial_stock > 0:
                    StockMovement.objects.create(
                        inventory=inventory,
                        unit=unit,
                        movement_type='IN',
                        quantity=initial_stock,
                        note='كمية ابتدائية عند إضافة وحدة جديدة للمنتج',
                        created_by=request.user,
                    )

            messages.success(request, f'تم تعديل المنتج "{product.name_ar}" بنجاح.')
            return redirect('staff:product_list')
    else:
        form = ProductForm(instance=product)
        formset = ProductUnitFormSet(instance=product)
    return render(request, 'staff/products/form.html', {
        'form': form,
        'formset': formset,
        'title': f'تعديل: {product.name_ar}',
        'is_edit': True,
        'product': product,
    })


@staff_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)

    has_stock = hasattr(product, 'inventory') and product.inventory.quantity > 0

    if request.method == 'POST':
        name = product.name_ar
        if has_stock:
            product.is_active = False
            product.save()
            messages.warning(request, f'المنتج "{name}" له مخزون — تم تعطيله بدل الحذف.')
        else:
            product.delete()
            messages.success(request, f'تم حذف المنتج "{name}".')
        return redirect('staff:product_list')

    return render(request, 'staff/products/delete.html', {
        'product': product,
        'has_stock': has_stock,
    })


@staff_required
def import_products(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'يرجى اختيار ملف Excel أولاً.')
            return redirect('staff:import_products')
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'يجب أن يكون الملف بصيغة .xlsx')
            return redirect('staff:import_products')
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            # الأعمدة السبعة كلها مطلوبة دلوقتي. سعر القطعة (الوحدة الصغرى)
            # مبقاش بيتكتب في الملف خالص — بيتحسب دايمًا تلقائيًا من
            # large_unit_price ÷ large_qty_in_small.
            required_headers = [
                'name_ar', 'category_slug',
                'large_unit_name', 'large_qty_in_small', 'large_unit_price',
                'unit_name', 'initial_stock',
            ]
            missing = [h for h in required_headers if h not in headers]
            if missing:
                messages.error(request, f'الأعمدة التالية ناقصة في الملف: {", ".join(missing)}')
                return redirect('staff:import_products')
            idx = {h: headers.index(h) for h in headers if h}
            success_count = 0
            error_rows = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    name_ar = str(row[idx['name_ar']]).strip() if row[idx['name_ar']] else ''
                    category_slug = str(row[idx['category_slug']]).strip() if row[idx['category_slug']] else ''
                    unit_name = str(row[idx['unit_name']]).strip() if row[idx['unit_name']] else ''
                    initial_stock = int(row[idx['initial_stock']]) if row[idx['initial_stock']] else 0

                    # --- بيانات الوحدة الكبرى (كرتونة) — مطلوبة لكل سطر ---
                    large_name = str(row[idx['large_unit_name']]).strip() if row[idx['large_unit_name']] else ''
                    raw_large_qty = row[idx['large_qty_in_small']]
                    raw_large_price = row[idx['large_unit_price']]
                    large_qty = int(raw_large_qty) if raw_large_qty else None
                    large_price = float(raw_large_price) if raw_large_price else None

                    if not name_ar or not category_slug or not unit_name \
                            or not large_name or not large_qty or not large_price:
                        error_rows.append(f'سطر {row_num}: بيانات ناقصة')
                        continue

                    # --- سعر القطعة (الوحدة الصغرى): محسوب تلقائيًا دايمًا ---
                    unit_price = round(large_price / large_qty, 2)

                    try:
                        category = Category.objects.get(slug=category_slug)
                    except Category.DoesNotExist:
                        error_rows.append(f'سطر {row_num}: القسم "{category_slug}" مش موجود')
                        continue
                    product, created = Product.objects.get_or_create(
                        name_ar=name_ar,
                        defaults={
                            'category': category,
                            'is_active': True,
                        }
                    )
                    if not created:
                        product.category = category
                        product.save()
                    unit, _ = ProductUnit.objects.update_or_create(
                        product=product,
                        size='S',
                        defaults={
                            'name': unit_name,
                            'unit_price': unit_price,
                            'qty_in_small': 1,
                        }
                    )
                    ProductUnit.objects.update_or_create(
                        product=product,
                        size='L',
                        defaults={
                            'name': large_name,
                            'unit_price': large_price,
                            'qty_in_small': large_qty,
                        }
                    )
                    inventory, _ = Inventory.objects.get_or_create(
                        product=product,
                        defaults={'quantity': 0, 'reserved': 0, 'min_quantity': 0}
                    )
                    if initial_stock > 0:
                        StockMovement.objects.create(
                            inventory=inventory,
                            unit=unit,
                            movement_type='IN',
                            quantity=initial_stock,
                            note='إضافة من ملف Excel',
                            created_by=request.user,
                        )
                    success_count += 1
                except Exception as e:
                    error_rows.append(f'سطر {row_num}: خطأ — {str(e)}')
                    continue
            if success_count:
                messages.success(request, f'تم إضافة/تحديث {success_count} منتج بنجاح.')
            for err in error_rows:
                messages.warning(request, err)
        except Exception as e:
            messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
    return render(request, 'staff/products/import.html')


@staff_required
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    headers = [
        'name_ar', 'category_slug',
        'large_unit_name', 'large_qty_in_small', 'large_unit_price',
        'unit_name', 'initial_stock',
    ]
    ws.append(headers)
    # سعر القطعة = large_unit_price ÷ large_qty_in_small تلقائيًا،
    # initial_stock دايمًا بالقطعة (الوحدة الصغرى).
    # مثال 1: كرتونة فيها 50 قطعة بسعر 100 جنيه -> سعر القطعة = 2.00 جنيه
    ws.append(['شاش طبي', 'gauze', 'كرتونة', 50, 100.00, 'قطعة', 200])
    # مثال 2: كرتونة فيها 10 علب بسعر 250 جنيه -> سعر العلبة = 25.00 جنيه
    ws.append(['قفازات لاتكس', 'gloves', 'كرتونة', 10, 250.00, 'علبة', 100])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    from django.http import HttpResponse
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="biozone_products_template.xlsx"'
    return response
