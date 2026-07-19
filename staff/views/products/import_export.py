"""
استيراد/تصدير المنتجات من وإلى ملفات إكسل. منطق CRUD الأساسي (عرض/إضافة/
تعديل/حذف) منفصل في crud.py — راجع staff/views/products/__init__.py
للتوثيق الكامل لسبب الفصل.
"""

from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from accounts.models import AccountType
from products.models import Product, ProductUnit, Category, UnitDiscount
from products.matching import normalize_name, find_similar_products
from products.new_arrivals import NEW_ARRIVALS_WINDOW_DAYS
from inventory.models import Inventory, StockMovement
from staff.permissions import perm_required
import openpyxl

IMPORT_SESSION_KEY = 'product_import_batch'
# حماية من ملف إكسل ضخم بالغلط (أو مقصود): الدفعة بالكامل بتتخزن مؤقتًا في
# الـ session (قاعدة البيانات) بين شاشة المراجعة وشاشة التأكيد، فملف بعشرات
# الآلاف من الصفوف كان بيعمل صف session ضخم ويشغل الـ worker وقت طويل في
# طلب واحد. الحدين دول سقف منطقي لأي استيراد حقيقي (لو المخزن عنده كتالوج
# أكبر فعلاً، يقسّم الملف على أكتر من دفعة).
IMPORT_MAX_FILE_SIZE_MB = 5
IMPORT_MAX_ROWS = 3000

FUZZY_MATCH_THRESHOLD = 0.82  # 82% تشابه فأكثر = "محتاج مراجعة بشرية"

# عمود الخصم لكل نوع حساب (فئة) بيتسمى discount:<اسم نوع الحساب> — الأنواع
# نفسها ديناميكية (بتتضاف/تتحذف من شاشة "أنواع الحسابات")، فمفيش عدد أعمدة
# ثابت: القالب/التصدير بيولّد عمود لكل نوع موجود وقت التحميل، والاستيراد
# بيدوّر على أي عمود بادئته discount: ويطابقه بالاسم مع الأنواع الحالية —
# لو النوع اتحذف أو الاسم اتغيّر، العمود بيتجاهل بدل ما يفشل الاستيراد كله.
DISCOUNT_COL_PREFIX = 'discount:'


def _discount_col_name(account_type):
    return f'{DISCOUNT_COL_PREFIX}{account_type.name}'


def _parse_unit_row(row_num, row, idx, account_types_by_col):
    """
    بيقرا صف واحد من شيت الإكسل — كل صف بيمثّل وحدة واحدة بس (قطعة أو
    كرتونة) لصنف معيّن، مش الصنف كامل. الصنف اللي له وحدتين (صغرى وكبرى)
    بيتكرر في صفين بنفس code (أو نفس الاسم) — راجع _group_unit_rows تحت
    اللي بتلمّهم مرة تانية في صنف واحد قبل الحفظ.
    """
    def cell(key):
        pos = idx.get(key)
        return row[pos] if pos is not None and pos < len(row) else None

    name_ar = str(cell('name_ar')).strip() if cell('name_ar') else ''
    category_slug = str(cell('category_slug')).strip() if cell('category_slug') else ''
    unit_name = str(cell('unit_name')).strip() if cell('unit_name') else ''
    code = str(cell('code')).strip() if cell('code') else ''

    raw_qty_in_small = cell('qty_in_small')
    raw_unit_price = cell('unit_price')
    raw_quantity = cell('quantity')

    if not name_ar or not unit_name or not raw_qty_in_small or not raw_unit_price:
        return None, f'سطر {row_num}: بيانات ناقصة (الاسم/الوحدة/الكمية بالقطعة/سعر الجمهور)'

    try:
        qty_in_small = int(raw_qty_in_small)
        unit_price = round(float(raw_unit_price), 2)
        quantity = int(raw_quantity) if raw_quantity else 0
    except (TypeError, ValueError):
        return None, f'سطر {row_num}: قيم رقمية غير صالحة'

    if qty_in_small < 1:
        return None, f'سطر {row_num}: "الكمية بالوحدة الصغرى" يجب أن تكون 1 على الأقل'

    if category_slug and not Category.objects.filter(slug=category_slug).exists():
        return None, f'سطر {row_num}: القسم "{category_slug}" مش موجود'

    # عمود discount:<فئة> موجود في الملف = القيمة دي هي الوصف الكامل لحالة
    # الخصم لهذا النوع، تمامًا زي شاشة "قائمة الخصومات" اليدوية: فاضي يعني
    # "امسح الخصم" مش "سيبه زي ما هو". نوع مالوش عمود أصلًا في الملف (مش
    # ضمن account_types_by_col) بيتسيب بدون أي تغيير.
    discounts = {}
    for col_name, account_type in account_types_by_col.items():
        raw = cell(col_name)
        if raw is None or str(raw).strip() == '':
            discounts[account_type.pk] = None  # إزالة صريحة لأي خصم موجود
            continue
        try:
            pct = Decimal(str(raw).strip())
        except InvalidOperation:
            return None, f'سطر {row_num}: نسبة خصم غير صالحة لـ"{account_type.name}"'
        if pct < 0 or pct > 100:
            return None, f'سطر {row_num}: نسبة الخصم لـ"{account_type.name}" يجب أن تكون بين 0 و100'
        discounts[account_type.pk] = str(pct)  # str عشان يفضل JSON-safe في الـ session

    return {
        'row_num': row_num,
        'code': code,
        'category_slug': category_slug,
        'name_ar': name_ar,
        'unit_name': unit_name,
        'qty_in_small': qty_in_small,
        'unit_price': unit_price,
        'quantity': quantity,
        'discounts': discounts,
    }, None


def _group_unit_rows(unit_rows):
    """
    بتجمع صفوف الوحدات (كل صف = وحدة واحدة) في صفوف "أصناف" — صنف بوحدة
    واحدة (كبرى بس مثلاً) أو بوحدتين (صغرى + كبرى) بيتعرف عن طريق نفس
    الـ code (لو موجود) أو نفس الاسم (بعد التطبيع) عبر كل صفوف الملف.
    نسبة الخصم بتتاخد من صف الوحدة الصغرى لو موجودة، وإلا من الوحدة
    الوحيدة (لو الصنف مالوش صغرى أصلًا) — نفس قاعدة التسعير في الموديل.
    """
    groups = {}
    order = []
    for ur in unit_rows:
        key = ('code', ur['code']) if ur['code'] else ('name', normalize_name(ur['name_ar']))
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(ur)

    products_data, errors = [], []
    for key in order:
        rows = groups[key]
        row_nums = ', '.join(str(r['row_num']) for r in rows)
        if len(rows) > 2:
            errors.append(f'صنف "{rows[0]["name_ar"]}": أكتر من وحدتين للصنف ده في الملف (سطور {row_nums})')
            continue
        small_rows = [r for r in rows if r['qty_in_small'] == 1]
        large_rows = [r for r in rows if r['qty_in_small'] > 1]
        if len(small_rows) > 1 or len(large_rows) > 1:
            errors.append(f'صنف "{rows[0]["name_ar"]}": فيه أكتر من صف بنفس حجم الوحدة (سطور {row_nums})')
            continue
        small = small_rows[0] if small_rows else None
        large = large_rows[0] if large_rows else None
        discount_source = small or large
        category_slug = next((r['category_slug'] for r in rows if r['category_slug']), '')
        code = next((r['code'] for r in rows if r['code']), '')
        products_data.append({
            'row_num': rows[0]['row_num'],
            'row_nums': [r['row_num'] for r in rows],
            'code': code,
            'category_slug': category_slug,
            'name_ar': rows[0]['name_ar'],
            'small': small,
            'large': large,
            'discounts': discount_source['discounts'] if discount_source else {},
        })
    return products_data, errors


def _classify_row(row_data, existing_by_code, existing_by_name_key, all_products):
    """
    بيحدد إيه اللي المفروض يحصل للصف ده اعتمادًا على مطابقة الكود (لو
    موجود) ثم الاسم المُطبَّع (بعد إزالة فروق المسافات/الأرقام/الحروف
    الشكلية) ثم أقرب الأسماء تشابهًا لو مفيش تطابق تام. النتيجة action
    واحدة من: update (تحديث صنف معروف بثقة) أو review (يحتاج قرار بشري)
    أو create (صنف جديد فعلًا، مفيش أي شبه بحاجة موجودة).
    """
    name_key = normalize_name(row_data['name_ar'])
    row_data['name_key'] = name_key

    if row_data['code'] and row_data['code'] in existing_by_code:
        product = existing_by_code[row_data['code']]
        return {**row_data, 'action': 'update', 'match_pk': product.pk,
                'match_name': product.name_ar, 'match_reason': 'code'}

    if name_key in existing_by_name_key:
        product = existing_by_name_key[name_key]
        return {**row_data, 'action': 'update', 'match_pk': product.pk,
                'match_name': product.name_ar, 'match_reason': 'name'}

    candidates = find_similar_products(name_key, all_products, threshold=FUZZY_MATCH_THRESHOLD)
    if candidates:
        return {**row_data, 'action': 'review', 'match_pk': None,
                'candidates': [{'pk': p.pk, 'name': p.name_ar, 'code': p.code, 'score': s}
                                for p, s in candidates]}

    return {**row_data, 'action': 'create', 'match_pk': None}


def _commit_product(row_data, target_pk, user, account_types_by_pk):
    """
    بيطبّق صنف واحد (وحدة أو وحدتين + خصوماته) فعليًا على قاعدة البيانات،
    بعد ما يبقى معروف بالظبط (من مرحلة المراجعة) هل ده تحديث لمنتج
    target_pk معين، ولا إضافة صنف جديد (target_pk=None). الكمية بتتسجل
    دايمًا كحركة "وارد" (IN) بتتضاف فوق الرصيد الحالي — مش استبدال له —
    سواء كانت "رصيد افتتاحي" لصنف جديد أو "تحديث كميات" لصنف موجود.
    بيرجّع (created, restocked).
    """
    category = None
    if row_data['category_slug']:
        category = Category.objects.get(slug=row_data['category_slug'])

    if target_pk:
        product = Product.objects.get(pk=target_pk)
        product.name_ar = row_data['name_ar']
        if category:
            product.category = category
        product.save()
        created = False
    else:
        if not category:
            raise ValueError(f'صنف جديد "{row_data["name_ar"]}" لازم يكون له قسم (category_slug)')
        product = Product.objects.create(name_ar=row_data['name_ar'], category=category, is_active=True)
        created = True

    inventory, _ = Inventory.objects.get_or_create(
        product=product, defaults={'quantity': 0, 'reserved': 0, 'min_quantity': 0},
    )

    restocked = False
    for size, unit_data in (('S', row_data['small']), ('L', row_data['large'])):
        if not unit_data:
            continue
        unit, _ = ProductUnit.objects.update_or_create(
            product=product, size=size,
            defaults={
                'name': unit_data['unit_name'],
                'unit_price': unit_data['unit_price'],
                'qty_in_small': unit_data['qty_in_small'],
            },
        )
        if unit_data['quantity'] > 0:
            StockMovement.objects.create(
                inventory=inventory, unit=unit, movement_type='IN',
                quantity=unit_data['quantity'], note='إضافة/تحديث من ملف Excel', created_by=user,
            )
            restocked = True

    # الخصم بيتحدد دايمًا على الوحدة "الأساسية" للتسعير: الصغرى لو موجودة
    # للصنف (حتى لو مكانتش في الملف ده تحديدًا، لأنها ممكن تكون اتضافت
    # قبل كده)، وإلا الوحدة الوحيدة المتاحة — راجع نفس القاعدة في
    # ProductUnit.get_pricing_breakdown_for_account_type.
    discount_unit = ProductUnit.objects.filter(product=product, size='S').first() \
        or ProductUnit.objects.filter(product=product, size='L').first()

    if discount_unit is not None:
        for at_pk_raw, pct_raw in row_data['discounts'].items():
            account_type = account_types_by_pk.get(int(at_pk_raw))
            if not account_type:
                continue
            if pct_raw is None:
                UnitDiscount.objects.filter(unit=discount_unit, account_type=account_type).delete()
            else:
                UnitDiscount.objects.update_or_create(
                    unit=discount_unit, account_type=account_type,
                    defaults={'discount_percent': Decimal(pct_raw)},
                )

    return created, restocked


@perm_required('products.add_product')
def import_products(request):
    """
    المرحلة الأولى: قراءة الملف وتصنيف كل صف (تحديث أكيد / إضافة جديدة
    أكيدة / يحتاج مراجعة) من غير أي حفظ فعلي، ثم عرض شاشة مراجعة. الحفظ
    الفعلي بيحصل بس في import_products_confirm بعد موافقة الموظف على أي
    صف يحتاج قرار بشري (اسم قريب من صنف موجود).
    """
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'يرجى اختيار ملف Excel أولاً.')
            return redirect('staff:import_products')
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'يجب أن يكون الملف بصيغة .xlsx')
            return redirect('staff:import_products')
        if excel_file.size > IMPORT_MAX_FILE_SIZE_MB * 1024 * 1024:
            messages.error(
                request,
                f'حجم الملف أكبر من الحد المسموح ({IMPORT_MAX_FILE_SIZE_MB} ميجا). '
                f'يرجى تقسيم الملف على أكتر من دفعة استيراد.'
            )
            return redirect('staff:import_products')
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            required_headers = ['name_ar', 'unit_name', 'qty_in_small', 'unit_price']
            missing = [h for h in required_headers if h not in headers]
            if missing:
                messages.error(request, f'الأعمدة التالية ناقصة في الملف: {", ".join(missing)}')
                return redirect('staff:import_products')
            idx = {h: headers.index(h) for h in headers if h}

            account_types = list(AccountType.objects.all().order_by('name'))
            account_types_by_col = {
                _discount_col_name(at): at for at in account_types if _discount_col_name(at) in idx
            }

            all_products = list(Product.objects.only('id', 'name_ar', 'code', 'name_key'))
            existing_by_code = {p.code: p for p in all_products if p.code}
            existing_by_name_key = {p.name_key: p for p in all_products if p.name_key}

            unit_rows, errors = [], []
            too_many_rows = False
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                if row_num - 1 > IMPORT_MAX_ROWS:
                    too_many_rows = True
                    break
                try:
                    row_data, error = _parse_unit_row(row_num, row, idx, account_types_by_col)
                except Exception as e:
                    error = f'سطر {row_num}: خطأ — {str(e)}'
                    row_data = None
                if error:
                    errors.append(error)
                    continue
                unit_rows.append(row_data)

            wb.close()

            if too_many_rows:
                messages.error(
                    request,
                    f'الملف فيه أكتر من {IMPORT_MAX_ROWS} صف. يرجى تقسيمه على أكتر من دفعة استيراد.'
                )
                return redirect('staff:import_products')

            products_data, group_errors = _group_unit_rows(unit_rows)
            errors.extend(group_errors)

            rows = [
                _classify_row(p, existing_by_code, existing_by_name_key, all_products)
                for p in products_data
            ]

            # صنف جديد (مش تحديث لصنف معروف) لازم يكون له قسم محدد في الملف —
            # صنف بيتحدّث (update) ممكن يسيب category_slug فاضي ويفضل على قسمه الحالي.
            valid_rows = []
            for r in rows:
                if r['action'] == 'create' and not r['category_slug']:
                    errors.append(f'سطر {r["row_num"]}: صنف جديد "{r["name_ar"]}" لازم يكون له قسم (category_slug)')
                    continue
                valid_rows.append(r)
            rows = valid_rows

            if not rows:
                messages.error(request, 'مفيش أي صف صالح في الملف.')
                for err in errors:
                    messages.warning(request, err)
                return redirect('staff:import_products')

            request.session[IMPORT_SESSION_KEY] = {'rows': rows, 'errors': errors}
            return redirect('staff:import_products_review')
        except Exception as e:
            messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
    return render(request, 'staff/products/import.html')


@perm_required('products.add_product')
def import_products_review(request):
    """
    شاشة المراجعة: بتعرض عدد الأصناف اللي هتتحدّث/هتتضاف بثقة تلقائيًا،
    وبتوقف عند أي صف اسمه قريب من صنف موجود وتسأل الموظف صراحةً هل ده
    نفس الصنف (تحديث) ولا صنف جديد فعلًا — قبل أي حفظ في قاعدة البيانات.
    """
    batch = request.session.get(IMPORT_SESSION_KEY)
    if not batch:
        messages.error(request, 'مفيش عملية استيراد جارية. من فضلك ارفع الملف تاني.')
        return redirect('staff:import_products')

    rows = batch['rows']
    context = {
        'errors': batch['errors'],
        'update_rows': [r for r in rows if r['action'] == 'update'],
        'create_rows': [r for r in rows if r['action'] == 'create'],
        'review_rows': [r for r in rows if r['action'] == 'review'],
        'new_arrivals_window_days': NEW_ARRIVALS_WINDOW_DAYS,
    }
    return render(request, 'staff/products/import_review.html', context)


@perm_required('products.add_product')
def import_products_confirm(request):
    """
    المرحلة التانية: بتاخد قرارات الموظف على صفوف "المراجعة" (اتحدد لكل
    واحد منها إما تحديث صنف بعينه أو إضافته كصنف جديد فعلًا) وتنفّذ الحفظ
    الفعلي لكل صفوف الدفعة مرة واحدة داخل transaction واحدة.
    """
    if request.method != 'POST':
        return redirect('staff:import_products')

    batch = request.session.get(IMPORT_SESSION_KEY)
    if not batch:
        messages.error(request, 'انتهت صلاحية عملية الاستيراد دي. من فضلك ارفع الملف تاني.')
        return redirect('staff:import_products')

    rows = batch['rows']
    account_types_by_pk = {at.pk: at for at in AccountType.objects.all()}
    created_count = updated_count = restocked_count = 0
    try:
        with transaction.atomic():
            for row_data in rows:
                if row_data['action'] == 'review':
                    decision = request.POST.get(f"decision_{row_data['row_num']}", 'new')
                    target_pk = int(decision) if decision != 'new' else None
                else:
                    target_pk = row_data.get('match_pk')
                created, restocked = _commit_product(row_data, target_pk, request.user, account_types_by_pk)
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    if restocked:
                        restocked_count += 1
    except Exception as e:
        messages.error(request, f'حصل خطأ أثناء الحفظ ولم يتم حفظ أي صنف: {str(e)}')
        return redirect('staff:import_products')

    del request.session[IMPORT_SESSION_KEY]
    if created_count:
        messages.success(request, f'تم إضافة {created_count} صنف جديد.')
    if updated_count:
        messages.success(request, f'تم تحديث {updated_count} صنف موجود.')
    for err in batch['errors']:
        messages.warning(request, err)

    # إشعار العملاء بالوارد الجديد — اختياري، الموظف بيحدده من شاشة المراجعة.
    # new_arrival_at اتحدّث تلقائيًا لكل صنف جديد أو اتزوّد رصيده (راجع
    # Product.save و StockMovement.save)، فمفيش داعي نحسب حاجة تانية هنا —
    # بس نتأكد إن فيه فعلاً حاجة جديدة تستاهل إشعار قبل ما نبعته.
    new_arrivals_total = created_count + restocked_count
    if request.POST.get('notify_clients') == 'on' and new_arrivals_total > 0:
        from notifications.services import notify_all_clients
        notify_all_clients(
            kind='NEW_ARRIVALS',
            title='وارد جديد في المتجر 🆕',
            message=f'تم إضافة {new_arrivals_total} صنف جديد أو تزويد رصيده — اطّلع على صفحة الوارد.',
            url_name='store:new_arrivals',
        )
        messages.success(request, 'تم إرسال إشعار الوارد الجديد لكل العملاء.')

    return redirect('staff:product_list')


@perm_required('products.view_product')
def download_template(request):
    """
    قالب فيه صف لكل وحدة (مش صف لكل صنف): الصنف اللي له وحدتين (قطعة +
    كرتونة) بيتكرر في صفين بنفس code (أو نفس الاسم بالظبط)، والصنف اللي
    له وحدة واحدة بس بياخد صف واحد. عمود discount:<اسم نوع الحساب> بيتولّد
    تلقائيًا لكل نوع حساب موجود دلوقتي — تقدر تضيف/تشيل عمود فئة كامل من
    الملف حسب احتياجك (لو النوع مش موجود وقت الرفع، العمود هيتجاهل).
    """
    account_types = list(AccountType.objects.all().order_by('name'))
    discount_headers = [_discount_col_name(at) for at in account_types]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    headers = [
        'code', 'category_slug', 'name_ar', 'unit_name',
        'qty_in_small', 'unit_price', 'quantity',
    ] + discount_headers
    ws.append(headers)

    blank_discounts = ['' for _ in account_types]
    small_discounts = [10 for _ in account_types]  # مثال: 10% لكل الفئات على القطعة
    large_discounts = [15 for _ in account_types]  # مثال: صنف بوحدة واحدة (كبرى بس)

    # مثال 1: صنف بوحدتين — الخصم بيتكتب على صف الوحدة الصغرى بس (قطعة)،
    # وصف الكرتونة بيتسيب فاضي لأن سعرها بيتحسب تلقائيًا من نسبة القطعة.
    ws.append(['', 'gauze', 'شاش طبي', 'قطعة', 1, 2.00, 200] + small_discounts)
    ws.append(['', 'gauze', 'شاش طبي', 'كرتونة', 50, 100.00, 0] + blank_discounts)

    # مثال 2: صنف بوحدة واحدة بس (كبرى) — الخصم بيتكتب على صفها هي نفسها.
    ws.append(['', 'gloves', 'قفازات لاتكس', 'كرتونة', 10, 250.00, 100] + large_discounts)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    from django.http import HttpResponse
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="biozone_products_template.xlsx"'
    return response


def _build_products_export_workbook(products):
    """
    بتبني ملف إكسل بنفس أعمدة قالب الاستيراد (صف لكل وحدة) لأي مجموعة
    أصناف (كل الأصناف، أو مجموعة مُنتقاة بالبحث/القسم) — مستخدمة في
    export_products (تصدير الكل) وexport_products_selected (تصدير المحدد).
    عمود code معبّى بكود كل صنف وأعمدة discount:<فئة> معبّية بنسبة الخصم
    الحالية لكل نوع حساب، عشان لو رفعت الملف تاني بعد التعديل، النظام
    يتعرّف على كل صنف بكوده ويحدّثه بدل ما يضيفه كصنف جديد. عمود quantity
    بيتصدّر دايمًا صفر (كمية "وارد" هتتضاف فوق الرصيد الحالي، مش الرصيد نفسه).
    """
    account_types = list(AccountType.objects.all().order_by('name'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    headers = [
        'code', 'category_slug', 'name_ar', 'unit_name',
        'qty_in_small', 'unit_price', 'quantity',
    ] + [_discount_col_name(at) for at in account_types]
    ws.append(headers)

    for product in products:
        units = list(product.units.all())
        small = next((u for u in units if u.size == 'S'), None)
        large = next((u for u in units if u.size == 'L'), None)
        discount_unit = small or large
        discount_by_pk = {}
        if discount_unit:
            discount_by_pk = {d.account_type_id: d.discount_percent for d in discount_unit.discounts.all()}
        discount_cells = [
            float(discount_by_pk[at.pk]) if at.pk in discount_by_pk else '' for at in account_types
        ]
        blank_discounts = ['' for _ in account_types]

        if small:
            ws.append([
                product.code, product.category.slug, product.name_ar, small.name,
                1, float(small.unit_price), 0,
            ] + discount_cells)
        if large:
            ws.append([
                product.code, product.category.slug, product.name_ar, large.name,
                large.qty_in_small, float(large.unit_price), 0,
            ] + (blank_discounts if small else discount_cells))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return wb


def _products_xlsx_response(wb, filename):
    from django.http import HttpResponse
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@perm_required('products.view_product')
def export_products(request):
    """تصدير كل الأصناف الحالية دفعة واحدة (بدون اختيار)."""
    products = Product.objects.select_related('category').prefetch_related(
        'units__discounts__account_type',
    ).all()
    wb = _build_products_export_workbook(products)
    return _products_xlsx_response(wb, 'biozone_products_export.xlsx')


@perm_required('products.view_product')
def export_products_select(request):
    """
    صفحة اختيار الأصناف قبل التصدير: تقدر تبحث وتحدد أصناف بعينها، أو
    تحدد قسم كامل (وبعدين تشيل منه أي صنف مش عايزه)، والنظام هيصدّر بس
    اللي محدد فعليًا لملف إكسل بنفس صيغة "تصدير الأصناف الحالية".
    """
    categories = Category.objects.filter(is_active=True).order_by('name')
    products = Product.objects.select_related('category').order_by('name_ar')
    products_data = [
        {
            'id': p.pk,
            'name': p.name_ar,
            'code': p.code,
            'category_slug': p.category.slug,
            'category_name': p.category.name,
        }
        for p in products
    ]
    return render(request, 'staff/products/export_select.html', {
        'categories': categories,
        'products_json': products_data,
    })


@perm_required('products.view_product')
def export_products_selected(request):
    """يستقبل قائمة IDs من صفحة الاختيار ويصدّرها كملف إكسل واحد."""
    if request.method != 'POST':
        return redirect('staff:export_products_select')

    ids = [pk for pk in request.POST.getlist('product_ids') if pk.isdigit()]
    if not ids:
        messages.warning(request, 'لازم تحدد صنف واحد على الأقل قبل التصدير.')
        return redirect('staff:export_products_select')

    products = Product.objects.select_related('category').prefetch_related(
        'units__discounts__account_type',
    ).filter(pk__in=ids)
    wb = _build_products_export_workbook(products)
    return _products_xlsx_response(wb, 'biozone_products_export_selected.xlsx')
