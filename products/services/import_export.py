"""
منطق استيراد/تصدير المنتجات من وإلى ملفات Excel.

كل الدوال هنا "نقية" قدر الإمكان: بتاخد بيانات وترجع بيانات، من غير ما
تعرف حاجة عن HTTP request أو response أو session. ده اللي بيخليها سهلة
الاختبار (products/services/tests.py) من غير ما نحتاج نمرّ بـ Django
test client أو نرفع ملف Excel فعلي لكل اختبار.

المسؤولية اتقسمت هنا لمرحلتين متطابقتين مع واجهة المستخدم:
1. القراءة والتصنيف (بدون أي حفظ): read_import_workbook + classify_row
2. الحفظ الفعلي بعد موافقة الموظف: commit_import_batch
"""
from decimal import Decimal, InvalidOperation

import openpyxl

from accounts.models import AccountType
from inventory.models import Inventory, StockMovement
from products.matching import normalize_name, find_similar_products
from products.models import Product, ProductUnit, Category, UnitDiscount

FUZZY_MATCH_THRESHOLD = 0.82  # 82% تشابه فأكثر = "محتاج مراجعة بشرية"

# عمود الخصم لكل نوع حساب (فئة) بيتسمى discount:<اسم نوع الحساب> — الأنواع
# نفسها ديناميكية (بتتضاف/تتحذف من شاشة "أنواع الحسابات")، فمفيش عدد أعمدة
# ثابت: القالب/التصدير بيولّد عمود لكل نوع موجود وقت التحميل، والاستيراد
# بيدوّر على أي عمود بادئته discount: ويطابقه بالاسم مع الأنواع الحالية —
# لو النوع اتحذف أو الاسم اتغيّر، العمود بيتجاهل بدل ما يفشل الاستيراد كله.
DISCOUNT_COL_PREFIX = 'discount:'

REQUIRED_IMPORT_HEADERS = ['name_ar', 'unit_name', 'qty_in_small', 'unit_price']


def discount_col_name(account_type):
    return f'{DISCOUNT_COL_PREFIX}{account_type.name}'


def parse_unit_row(row_num, row, idx, account_types_by_col):
    """
    بيقرا صف واحد من شيت الإكسل — كل صف بيمثّل وحدة واحدة بس (قطعة أو
    كرتونة) لصنف معيّن، مش الصنف كامل. الصنف اللي له وحدتين (صغرى وكبرى)
    بيتكرر في صفين بنفس code (أو نفس الاسم) — راجع group_unit_rows تحت
    اللي بتلمّهم مرة تانية في صنف واحد قبل الحفظ.
    """
    def cell(key):
        pos = idx.get(key)
        return row[pos] if pos is not None and pos < len(row) else None

    name_ar = str(cell('name_ar')).strip() if cell('name_ar') else ''
    category_slug = str(cell('category_slug')).strip() if cell('category_slug') else ''
    unit_name = str(cell('unit_name')).strip() if cell('unit_name') else ''
    code = str(cell('code')).strip() if cell('code') else ''

    raw_qty_in_small = cell('qty_in_small')
    raw_unit_price = cell('unit_price')
    raw_quantity = cell('quantity')

    if not name_ar or not unit_name or not raw_qty_in_small or not raw_unit_price:
        return None, f'سطر {row_num}: بيانات ناقصة (الاسم/الوحدة/الكمية بالقطعة/سعر الجمهور)'

    try:
        qty_in_small = int(raw_qty_in_small)
        unit_price = round(float(raw_unit_price), 2)
        quantity = int(raw_quantity) if raw_quantity else 0
    except (TypeError, ValueError):
        return None, f'سطر {row_num}: قيم رقمية غير صالحة'

    if qty_in_small < 1:
        return None, f'سطر {row_num}: "الكمية بالوحدة الصغرى" يجب أن تكون 1 على الأقل'

    if category_slug and not Category.objects.filter(slug=category_slug).exists():
        return None, f'سطر {row_num}: القسم "{category_slug}" مش موجود'

    # عمود discount:<فئة> موجود في الملف = القيمة دي هي الوصف الكامل لحالة
    # الخصم لهذا النوع، تمامًا زي شاشة "قائمة الخصومات" اليدوية: فاضي يعني
    # "امسح الخصم" مش "سيبه زي ما هو". نوع مالوش عمود أصلًا في الملف (مش
    # ضمن account_types_by_col) بيتسيب بدون أي تغيير.
    discounts = {}
    for col_name, account_type in account_types_by_col.items():
        raw = cell(col_name)
        if raw is None or str(raw).strip() == '':
            discounts[account_type.pk] = None  # إزالة صريحة لأي خصم موجود
            continue
        try:
            pct = Decimal(str(raw).strip())
        except InvalidOperation:
            return None, f'سطر {row_num}: نسبة خصم غير صالحة لـ"{account_type.name}"'
        if pct < 0 or pct > 100:
            return None, f'سطر {row_num}: نسبة الخصم لـ"{account_type.name}" يجب أن تكون بين 0 و100'
        discounts[account_type.pk] = str(pct)  # str عشان يفضل JSON-safe في الـ session

    return {
        'row_num': row_num,
        'code': code,
        'category_slug': category_slug,
        'name_ar': name_ar,
        'unit_name': unit_name,
        'qty_in_small': qty_in_small,
        'unit_price': unit_price,
        'quantity': quantity,
        'discounts': discounts,
    }, None


def group_unit_rows(unit_rows):
    """
    بتجمع صفوف الوحدات (كل صف = وحدة واحدة) في صفوف "أصناف" — صنف بوحدة
    واحدة (كبرى بس مثلاً) أو بوحدتين (صغرى + كبرى) بيتعرف عن طريق نفس
    الـ code (لو موجود) أو نفس الاسم (بعد التطبيع) عبر كل صفوف الملف.
    نسبة الخصم بتتاخد من صف الوحدة الصغرى لو موجودة، وإلا من الوحدة
    الوحيدة (لو الصنف مالوش صغرى أصلًا) — نفس قاعدة التسعير في الموديل.
    """
    groups = {}
    order = []
    for ur in unit_rows:
        key = ('code', ur['code']) if ur['code'] else ('name', normalize_name(ur['name_ar']))
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(ur)

    products_data, errors = [], []
    for key in order:
        rows = groups[key]
        row_nums = ', '.join(str(r['row_num']) for r in rows)
        if len(rows) > 2:
            errors.append(f'صنف "{rows[0]["name_ar"]}": أكتر من وحدتين للصنف ده في الملف (سطور {row_nums})')
            continue
        small_rows = [r for r in rows if r['qty_in_small'] == 1]
        large_rows = [r for r in rows if r['qty_in_small'] > 1]
        if len(small_rows) > 1 or len(large_rows) > 1:
            errors.append(f'صنف "{rows[0]["name_ar"]}": فيه أكتر من صف بنفس حجم الوحدة (سطور {row_nums})')
            continue
        small = small_rows[0] if small_rows else None
        large = large_rows[0] if large_rows else None
        discount_source = small or large
        category_slug = next((r['category_slug'] for r in rows if r['category_slug']), '')
        code = next((r['code'] for r in rows if r['code']), '')
        products_data.append({
            'row_num': rows[0]['row_num'],
            'row_nums': [r['row_num'] for r in rows],
            'code': code,
            'category_slug': category_slug,
            'name_ar': rows[0]['name_ar'],
            'small': small,
            'large': large,
            'discounts': discount_source['discounts'] if discount_source else {},
        })
    return products_data, errors


def classify_row(row_data, existing_by_code, existing_by_name_key, all_products):
    """
    بيحدد إيه اللي المفروض يحصل للصف ده اعتمادًا على مطابقة الكود (لو
    موجود) ثم الاسم المُطبَّع (بعد إزالة فروق المسافات/الأرقام/الحروف
    الشكلية) ثم أقرب الأسماء تشابهًا لو مفيش تطابق تام. النتيجة action
    واحدة من: update (تحديث صنف معروف بثقة) أو review (يحتاج قرار بشري)
    أو create (صنف جديد فعلًا، مفيش أي شبه بحاجة موجودة).
    """
    name_key = normalize_name(row_data['name_ar'])
    row_data['name_key'] = name_key

    if row_data['code'] and row_data['code'] in existing_by_code:
        product = existing_by_code[row_data['code']]
        return {**row_data, 'action': 'update', 'match_pk': product.pk,
                'match_name': product.name_ar, 'match_reason': 'code'}

    if name_key in existing_by_name_key:
        product = existing_by_name_key[name_key]
        return {**row_data, 'action': 'update', 'match_pk': product.pk,
                'match_name': product.name_ar, 'match_reason': 'name'}

    candidates = find_similar_products(name_key, all_products, threshold=FUZZY_MATCH_THRESHOLD)
    if candidates:
        return {**row_data, 'action': 'review', 'match_pk': None,
                'candidates': [{'pk': p.pk, 'name': p.name_ar, 'code': p.code, 'score': s}
                                for p, s in candidates]}

    return {**row_data, 'action': 'create', 'match_pk': None}


def read_import_workbook(excel_file, max_rows):
    """
    بتقرا ملف Excel كامل وترجّع (rows, errors, error_message).
    error_message لو موجودة معناها فشل عام (ملف غير صالح، أعمدة ناقصة،
    عدد صفوف أكبر من الحد المسموح) والعملية لازم توقف فورًا. غير كده،
    rows هي قائمة الأصناف المصنّفة (update/create/review) وerrors تحذيرات
    على مستوى صف واحد بس (باقي الصفوف الصحيحة اتعالجت عادي).
    """
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        missing = [h for h in REQUIRED_IMPORT_HEADERS if h not in headers]
        if missing:
            return [], [], f'الأعمدة التالية ناقصة في الملف: {", ".join(missing)}'
        idx = {h: headers.index(h) for h in headers if h}

        account_types = list(AccountType.objects.all().order_by('name'))
        account_types_by_col = {
            discount_col_name(at): at for at in account_types if discount_col_name(at) in idx
        }

        all_products = list(Product.objects.only('id', 'name_ar', 'code', 'name_key'))
        existing_by_code = {p.code: p for p in all_products if p.code}
        existing_by_name_key = {p.name_key: p for p in all_products if p.name_key}

        unit_rows, errors = [], []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            if row_num - 1 > max_rows:
                wb.close()
                return [], [], f'الملف فيه أكتر من {max_rows} صف. يرجى تقسيمه على أكتر من دفعة استيراد.'
            try:
                row_data, error = parse_unit_row(row_num, row, idx, account_types_by_col)
            except Exception as e:
                error = f'سطر {row_num}: خطأ — {str(e)}'
                row_data = None
            if error:
                errors.append(error)
                continue
            unit_rows.append(row_data)

        wb.close()
    except Exception as e:
        return [], [], f'خطأ في قراءة الملف: {str(e)}'

    products_data, group_errors = group_unit_rows(unit_rows)
    errors.extend(group_errors)

    rows = [
        classify_row(p, existing_by_code, existing_by_name_key, all_products)
        for p in products_data
    ]

    # صنف جديد (مش تحديث لصنف معروف) لازم يكون له قسم محدد في الملف —
    # صنف بيتحدّث (update) ممكن يسيب category_slug فاضي ويفضل على قسمه الحالي.
    valid_rows = []
    for r in rows:
        if r['action'] == 'create' and not r['category_slug']:
            errors.append(f'سطر {r["row_num"]}: صنف جديد "{r["name_ar"]}" لازم يكون له قسم (category_slug)')
            continue
        valid_rows.append(r)

    return valid_rows, errors, None


def commit_product(row_data, target_pk, user, account_types_by_pk):
    """
    بيطبّق صنف واحد (وحدة أو وحدتين + خصوماته) فعليًا على قاعدة البيانات،
    بعد ما يبقى معروف بالظبط (من مرحلة المراجعة) هل ده تحديث لمنتج
    target_pk معين، ولا إضافة صنف جديد (target_pk=None). الكمية بتتسجل
    دايمًا كحركة "وارد" (IN) بتتضاف فوق الرصيد الحالي — مش استبدال له —
    سواء كانت "رصيد افتتاحي" لصنف جديد أو "تحديث كميات" لصنف موجود.
    بيرجّع (created, restocked).
    """
    category = None
    if row_data['category_slug']:
        category = Category.objects.get(slug=row_data['category_slug'])

    if target_pk:
        product = Product.objects.get(pk=target_pk)
        product.name_ar = row_data['name_ar']
        if category:
            product.category = category
        product.save()
        created = False
    else:
        if not category:
            raise ValueError(f'صنف جديد "{row_data["name_ar"]}" لازم يكون له قسم (category_slug)')
        product = Product.objects.create(name_ar=row_data['name_ar'], category=category, is_active=True)
        created = True

    inventory, _ = Inventory.objects.get_or_create(
        product=product, defaults={'quantity': 0, 'reserved': 0, 'min_quantity': 0},
    )

    restocked = False
    for size, unit_data in (('S', row_data['small']), ('L', row_data['large'])):
        if not unit_data:
            continue
        unit, _ = ProductUnit.objects.update_or_create(
            product=product, size=size,
            defaults={
                'name': unit_data['unit_name'],
                'unit_price': unit_data['unit_price'],
                'qty_in_small': unit_data['qty_in_small'],
            },
        )
        if unit_data['quantity'] > 0:
            StockMovement.objects.create(
                inventory=inventory, unit=unit, movement_type='IN',
                quantity=unit_data['quantity'], note='إضافة/تحديث من ملف Excel', created_by=user,
            )
            restocked = True

    # الخصم بيتحدد دايمًا على الوحدة "الأساسية" للتسعير: الصغرى لو موجودة
    # للصنف (حتى لو مكانتش في الملف ده تحديدًا، لأنها ممكن تكون اتضافت
    # قبل كده)، وإلا الوحدة الوحيدة المتاحة — راجع نفس القاعدة في
    # ProductUnit.get_pricing_breakdown_for_account_type.
    discount_unit = ProductUnit.objects.filter(product=product, size='S').first() \
        or ProductUnit.objects.filter(product=product, size='L').first()

    if discount_unit is not None:
        for at_pk_raw, pct_raw in row_data['discounts'].items():
            account_type = account_types_by_pk.get(int(at_pk_raw))
            if not account_type:
                continue
            if pct_raw is None:
                UnitDiscount.objects.filter(unit=discount_unit, account_type=account_type).delete()
            else:
                UnitDiscount.objects.update_or_create(
                    unit=discount_unit, account_type=account_type,
                    defaults={'discount_percent': Decimal(pct_raw)},
                )

    return created, restocked


def commit_import_batch(rows, decisions, user):
    """
    بتاخد قرارات الموظف على صفوف "المراجعة" (decisions: dict بمفتاح
    row_num وقيمة إما 'new' أو pk المنتج المستهدف) وتنفّذ الحفظ الفعلي
    لكل صفوف الدفعة. الـ transaction بيتحكم فيها المستدعي (الـ view) عشان
    تفضل الدالة دي قابلة لإعادة الاستخدام برّة سياق request لو احتجنا.
    بترجّع (created_count, updated_count, restocked_count).
    """
    account_types_by_pk = {at.pk: at for at in AccountType.objects.all()}
    created_count = updated_count = restocked_count = 0
    for row_data in rows:
        if row_data['action'] == 'review':
            decision = decisions.get(row_data['row_num'], 'new')
            target_pk = int(decision) if decision != 'new' else None
        else:
            target_pk = row_data.get('match_pk')
        created, restocked = commit_product(row_data, target_pk, user, account_types_by_pk)
        if created:
            created_count += 1
        else:
            updated_count += 1
            if restocked:
                restocked_count += 1
    return created_count, updated_count, restocked_count


def build_products_export_workbook(products):
    """
    بتبني ملف إكسل بنفس أعمدة قالب الاستيراد (صف لكل وحدة) لأي مجموعة
    أصناف (كل الأصناف، أو مجموعة مُنتقاة بالبحث/القسم) — مستخدمة في
    export_products (تصدير الكل) وexport_products_selected (تصدير المحدد).
    عمود code معبّى بكود كل صنف وأعمدة discount:<فئة> معبّية بنسبة الخصم
    الحالية لكل نوع حساب، عشان لو رفعت الملف تاني بعد التعديل، النظام
    يتعرّف على كل صنف بكوده ويحدّثه بدل ما يضيفه كصنف جديد. عمود quantity
    بيتصدّر دايمًا صفر (كمية "وارد" هتتضاف فوق الرصيد الحالي، مش الرصيد نفسه).
    """
    account_types = list(AccountType.objects.all().order_by('name'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    headers = [
        'code', 'category_slug', 'name_ar', 'unit_name',
        'qty_in_small', 'unit_price', 'quantity',
    ] + [discount_col_name(at) for at in account_types]
    ws.append(headers)

    for product in products:
        units = list(product.units.all())
        small = next((u for u in units if u.size == 'S'), None)
        large = next((u for u in units if u.size == 'L'), None)
        discount_unit = small or large
        discount_by_pk = {}
        if discount_unit:
            discount_by_pk = {d.account_type_id: d.discount_percent for d in discount_unit.discounts.all()}
        discount_cells = [
            float(discount_by_pk[at.pk]) if at.pk in discount_by_pk else '' for at in account_types
        ]
        blank_discounts = ['' for _ in account_types]

        if small:
            ws.append([
                product.code, product.category.slug, product.name_ar, small.name,
                1, float(small.unit_price), 0,
            ] + discount_cells)
        if large:
            ws.append([
                product.code, product.category.slug, product.name_ar, large.name,
                large.qty_in_small, float(large.unit_price), 0,
            ] + (blank_discounts if small else discount_cells))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return wb


def build_import_template_workbook():
    """قالب فارغ (بأمثلة توضيحية) لأعمدة الاستيراد — نفس أعمدة التصدير بالظبط."""
    account_types = list(AccountType.objects.all().order_by('name'))
    discount_headers = [discount_col_name(at) for at in account_types]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    headers = [
        'code', 'category_slug', 'name_ar', 'unit_name',
        'qty_in_small', 'unit_price', 'quantity',
    ] + discount_headers
    ws.append(headers)

    blank_discounts = ['' for _ in account_types]
    small_discounts = [10 for _ in account_types]  # مثال: 10% لكل الفئات على القطعة
    large_discounts = [15 for _ in account_types]  # مثال: صنف بوحدة واحدة (كبرى بس)

    # مثال 1: صنف بوحدتين — الخصم بيتكتب على صف الوحدة الصغرى بس (قطعة)،
    # وصف الكرتونة بيتسيب فاضي لأن سعرها بيتحسب تلقائيًا من نسبة القطعة.
    ws.append(['', 'gauze', 'شاش طبي', 'قطعة', 1, 2.00, 200] + small_discounts)
    ws.append(['', 'gauze', 'شاش طبي', 'كرتونة', 50, 100.00, 0] + blank_discounts)

    # مثال 2: صنف بوحدة واحدة بس (كبرى) — الخصم بيتكتب على صفها هي نفسها.
    ws.append(['', 'gloves', 'قفازات لاتكس', 'كرتونة', 10, 250.00, 100] + large_discounts)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return wb
